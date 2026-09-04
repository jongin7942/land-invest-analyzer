"""산업단지 신규지정 이력을 모은다 — 공공데이터포털의 분기별 과거 파일에서.

── 왜 ─────────────────────────────────────────────────────────────
'전국산업단지현황통계' 통합본에는 단지별 지정일자가 없다. 있는 건 부록1
'신규지정 및 해제현황' 인데 **그 분기 것만** 들어 있다(2025-4분기 24행).
그런데 포털에는 같은 데이터셋의 분기별 과거 파일이 31개(2020-06 ~ 2025-09)
쌓여 있다. 각 파일의 부록1을 모으면 지정일자가 붙은 신규지정 사례가
수백 건이 된다. 그게 있어야 "산업단지가 지정되면 근처 아파트가 움직이나" 를
교통·재건축과 같은 자로 잴 수 있다.

── 경로 (포털 JS 를 그대로 따라간다) ─────────────────────────────
  1. POST /tcs/dss/selectHistAndCsvData.do  → 과거 파일 목록 HTML
     (data-public-pk = publicDataDetailPk 역할의 uddi)
  2. POST /tcs/dss/selectFileDataDownload.do → JSON(atchFileId, fileDetailSn)
  3. POST /cmm/cmm/check-limit.json          → needCaptcha 확인
  4. GET  /cmm/cmm/fileDownload.do?atchFileId=..&fileDetailSn=..  → xlsx

받은 파일은 logs/kicox/ 에 두고, 부록1 시트만 읽어 rules/kicox_designations.csv 로
합친다. 같은 단지가 여러 분기에 나오면(지정 후 변경 등) 가장 이른 지정일을 쓴다.
"""
from __future__ import annotations

import csv
import json
import re
import sys
import time
import urllib.parse
import urllib.request
from http.cookiejar import CookieJar
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
BASE = "https://www.data.go.kr"
PK = "3041272"
DETAIL_PK = "uddi:bfab8ec1-d98a-454f-8eb6-ed2f90073428"
PAGE = f"{BASE}/data/{PK}/fileData.do"
DIR = ROOT / "logs" / "kicox"
OUT = ROOT / "rules" / "kicox_designations.csv"

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/131.0 Safari/537.36")


def opener():
    op = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(CookieJar()))
    op.addheaders = [("User-Agent", UA), ("Referer", PAGE), ("X-Requested-With", "XMLHttpRequest")]
    op.open(PAGE, timeout=30).read()
    return op


def post(op, url, fields):
    req = urllib.request.Request(url, data=urllib.parse.urlencode(fields).encode())
    with op.open(req, timeout=60) as r:
        return r.read().decode("utf-8", "replace")


def history(op) -> list[tuple[str, str, str]]:
    """[(파일명, detailPk, histSn)] — 최신 파일(통합본)도 포함한다."""
    html = post(op, f"{BASE}/tcs/dss/selectHistAndCsvData.do",
                {"publicDataPk": PK, "publicDataDetailPk": DETAIL_PK})
    items = re.findall(
        r'data-public-pk="([^"]+)"\s+data-public-detail-sn="(\d+)">\s*([^<]+?)\s*</a>', html)
    out = [(name.strip(), pk, sn) for pk, sn, name in items]
    out.insert(0, ("한국산업단지공단_전국산업단지현황통계_20251231", DETAIL_PK, "1"))
    return out


def download(op, name: str, detail_pk: str) -> Path | None:
    dest = DIR / f"{name}.xlsx"
    if dest.exists() and dest.stat().st_size > 10_000:
        return dest
    j = json.loads(post(op, f"{BASE}/tcs/dss/selectFileDataDownload.do", {
        "publicDataDetailPk": detail_pk, "publicDataPk": PK, "atchFileId": "",
        "fileDetailSn": "1", "publicDataTyCode": "PR0051"}))
    if not j.get("status") or not j.get("atchFileId"):
        print(f"  건너뜀 {name}: {j.get('error') or 'atchFileId 없음'}")
        return None
    chk = json.loads(post(op, f"{BASE}/cmm/cmm/check-limit.json",
                          {"atchFileId": j["atchFileId"], "fileDetailSn": j["fileDetailSn"]}))
    if chk.get("needCaptcha"):
        print(f"  건너뜀 {name}: 포털이 캡차를 요구합니다 (잠시 뒤 다시)")
        return None
    url = (f"{BASE}/cmm/cmm/fileDownload.do?atchFileId={j['atchFileId']}"
           f"&fileDetailSn={j['fileDetailSn']}&dataNm={urllib.parse.quote(name)}")
    with op.open(url, timeout=180) as r:
        data = r.read()
    if data[:2] != b"PK":
        print(f"  건너뜀 {name}: xlsx 가 아닙니다 ({data[:20]!r})")
        return None
    dest.write_bytes(data)
    return dest


def _sheet_rows(path: Path):
    """부록1(신규지정) 시트의 행들. 2018~2019년 파일은 csv/hwp/xls(x) 를 묶은 zip 이라
    안에서 엑셀을 꺼내 읽는다. 옛 .xls 는 xlrd 로 읽는다."""
    import io
    import zipfile
    blob = path.read_bytes()
    z = zipfile.ZipFile(io.BytesIO(blob))
    if "[Content_Types].xml" not in z.namelist():
        inner = next((n for n in z.namelist() if n.lower().endswith(".xlsx")), None)
        if inner:
            blob = z.read(inner)
        else:
            inner = next((n for n in z.namelist() if n.lower().endswith(".xls")), None)
            if not inner:
                return None
            import xlrd
            book = xlrd.open_workbook(file_contents=z.read(inner))
            name = next((n for n in book.sheet_names() if "신규지정" in n), None)
            if not name:
                return None
            sh = book.sheet_by_name(name)
            return [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    sheet = next((n for n in wb.sheetnames if "신규지정" in n), None)
    if not sheet:
        return None
    return list(wb[sheet].iter_rows(values_only=True))


def designations(path: Path) -> list[dict]:
    """부록1 시트에서 (유형, 시도, 시군구, 단지명, 지정면적, 산업용지, 지정일자)."""
    try:
        raw = _sheet_rows(path)
    except Exception as e:  # noqa: BLE001
        print(f"  읽기 실패 {path.name}: {e}")
        return []
    if not raw:
        return []
    rows, header = [], None
    for r in raw:
        cells = [("" if c is None else str(c).strip()) for c in r]
        if header is None:
            if "지정일자" in cells and "단지명" in cells:
                header = {name: i for i, name in enumerate(cells)}
            continue
        try:
            day = cells[header["지정일자"]]
            name = cells[header["단지명"]]
        except (KeyError, IndexError):
            continue
        m = re.match(r"(\d{4})[-./]?(\d{2})[-./]?(\d{2})", day.replace(" ", ""))
        if not m or not name or "합계" in name:
            continue
        rows.append({
            "kind": cells[header.get("유형", 0)], "sido": cells[header.get("시도", 1)],
            "sigungu": cells[header.get("시군구", 2)], "name": name,
            "area_k_m2": cells[header.get("지정면적", 4)],
            "industry_k_m2": cells[header.get("산업용지", 5)],
            "designated": f"{m.group(1)}-{m.group(2)}-{m.group(3)}",
            "note": cells[header["비고"]] if "비고" in header and header["비고"] < len(cells) else "",
            "source_file": path.name,
        })
    return rows


def main() -> int:
    DIR.mkdir(parents=True, exist_ok=True)
    op = opener()
    files = history(op)
    print(f"과거 파일 {len(files)}개")
    got = []
    for name, pk, sn in files:
        p = download(op, name, pk)
        if p:
            got.append(p)
            print(f"  받음 {name}")
        time.sleep(0.5)

    merged: dict[tuple[str, str], dict] = {}
    for p in sorted(got):
        for row in designations(p):
            key = (row["sido"], row["name"])
            if key not in merged or row["designated"] < merged[key]["designated"]:
                merged[key] = row
    rows = sorted(merged.values(), key=lambda r: r["designated"])
    with OUT.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()) if rows else ["name"])
        w.writeheader()
        w.writerows(rows)
    cap = [r for r in rows if r["sido"] in ("서울", "경기", "인천")]
    print(f"\n신규지정 {len(rows)}건 (수도권 {len(cap)}건) → {OUT}")
    if rows:
        print(f"  지정일 범위 {rows[0]['designated']} ~ {rows[-1]['designated']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
